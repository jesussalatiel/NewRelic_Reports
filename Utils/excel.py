from openpyxl import Workbook
from openpyxl import load_workbook
from credentials import data_credentials

class Excel:
    
    workbook = Workbook() #Instance woorkbook
     
    def save(self):
        try:
            self.workbook.save(filename = data_credentials['excel_file_name']) #Save the file in the system
            return True #Return True if the file was successful save
        except:
            return False

    def createIncidentsReport(self, elements_list): 
        try:
            sheet = self.workbook.active
            sheet.append(['Monitor','Dia','Hora de Inicio','Duracion', 'Tipo de Alerta']) #Set the titles
            for row in elements_list: # Show the list 
                sheet.append(row) #Add the list to excel
            
            #Create the little table that contains the sum of elements by IB, CTF and BM 
            sheet.merge_cells('I3:J3')
            sheet.cell(row = 3, column = 9).value = 'IB'
            sheet["I4"] = 'Tipo de Alerta'
            sheet["J4"] = 'Incidencias'
            sheet["I5"] = 'Otras'
            sheet["J5"] = 0
            sheet["I6"] = 'Error rate > 5.0%'
            sheet["J6"] = 0
            sheet["I7"] = 'Apdex Score < 0.7'
            sheet["J7"] = '=COUNTIF(E:E, "Apdex_IB")' #Formula to know how times you get the registers of IB
            sheet["J8"] = '=COUNTIF(E:E, "Availability_IB")' #Formula to know how times you get the registers of IB
            sheet["I8"] = 'Ping Alerts'
            sheet.merge_cells('K3:L3')
            sheet.cell(row = 3, column = 11).value = 'CTF'
            sheet["K4"] = 'Tipo de Alerta'
            sheet["L4"] = 'Incidencias'
            sheet["K5"] = 'Otras'
            sheet["L5"] = 0
            sheet["K6"] = 'Error rate > 5.0%'
            sheet["L6"] = 0
            sheet["K7"] = 'Apdex Score < 0.7'
            sheet["L7"] = '=COUNTIF(E:E, "Apdex_CTF")' #Formula to know how times you get the registers of CTF
            sheet["L8"] = '=COUNTIF(E:E, "Availability_CTF")' #Formula to know how times you get the registers of CTF
            sheet["K8"] = 'Ping Alerts'
            sheet.merge_cells('M3:N3')
            sheet.cell(row = 3, column = 13).value = 'BM'
            sheet["M4"] = 'Tipo de Alerta'
            sheet["N4"] = 'Incidencias'
            sheet["M5"] = 'Otras'
            sheet["N5"] = 0
            sheet["M6"] = 'Error rate > 5.0%'
            sheet["N6"] = 0
            sheet["M7"] = 'Apdex Score < 0.7'
            sheet["N7"] = '=COUNTIF(E:E, "Apdex_BM")' #Formula to know how times you get the registers of BM
            sheet["N8"] = '=COUNTIF(E:E, "Availability_BM")' #Formula to know how times you get the registers of BM
            sheet["M8"] = 'Ping Alerts'
            
        except:
            import sys
            print("Something went wrong: {0}".format(sys.exc_info()[0]))

        finally:
            if self.save() == True:   
                return True #This method return True if all was good
            else:
                return False    